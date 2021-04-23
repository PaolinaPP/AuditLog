#!/usr/bin/python3

from openpyxl import load_workbook
from collections import Counter

class FileOperations:

    def __init__(self, file_path):
        self.file_path = file_path

    ####################################################
    #    function name: load_file                      #
    #    function parameters: None                     #
    #    function description: loading file as sheet   #
    ####################################################
    def load_file(self):
        sheet = None
        try:
            workbook = load_workbook(filename=self.file_path)
            sheet = workbook.active
        except InvalidFileException:
            print("Invalid file.")    
        except Exception:
            print("An exception occurred while loading file.")
        return sheet

    ####################################################
    #    function name: read_file                      #
    #    function parameters: string str_to_search_for #
    #                         sheet (document copy)    #
    #    function description: select only rows which  #
    #        contains a specific string                #
    ####################################################
    def read_data(self, str_to_search_for, sheet):
        lectures = []
        try:
            for i in range(2, sheet.max_row, 1): #from 2 because data starts from row 2
                if str_to_search_for in sheet.cell(row=i, column=2).value:
                    lectures.append(sheet.cell(row=i, column=2).value)
        except Exception:
            print("An exception occurred while reading data.")
        return lectures   
        
    

    