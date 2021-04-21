#!/usr/bin/python3

import unittest
from pathlib import Path
from openpyxl import load_workbook
import sys
sys.path.append('..')
import scripts.file_operations as fo

class TestFileOperations(unittest.TestCase):

    c_file_path = Path(__file__).with_name('info_table.xlsx')
    c_fo_obj = fo.FileOperations(c_file_path)
    c_sheet = load_workbook(filename=c_file_path).active

    def test_load_file(self):
        actual = TestFileOperations.c_fo_obj.load_file()
        self.assertEqual(TestFileOperations.c_sheet.max_row, actual.max_row)
        self.assertEqual(TestFileOperations.c_sheet.max_column, actual.max_column)

    @unittest.expectedFailure
    def test_load_file_test(self):
        actual = TestFileOperations.c_fo_obj.load_file()
        self.assertEqual(TestFileOperations.c_sheet.max_row-5, actual.max_row)
        self.assertEqual(TestFileOperations.c_sheet.max_column-5, actual.max_column)    

    def test_read_data(self):
        expected = ['File: Лекция 8: Език за заявки SPARQL', 'File: Лекция 1: Въведение в програмиране за семантичен уеб', 'File: Лекция 7: Проектиране на онтология с Protégé', 'File: Лекция 9: Програмиране за семантичен уеб', 'File: Лекция 1: Въведение в програмиране за семантичен уеб']
        actual = TestFileOperations.c_fo_obj.read_data("File: Лекция", TestFileOperations.c_sheet)
        self.assertEqual(expected, actual)

    @unittest.expectedFailure
    def test_read_data_fail(self):
        expected = ['File: Лекция 1: Въведение в програмиране за семантичен уеб', 'File: Лекция 7: Проектиране на онтология с Protégé', 'File: Лекция 9: Програмиране за семантичен уеб', 'File: Лекция 1: Въведение в програмиране за семантичен уеб']
        actual = TestFileOperations.c_fo_obj.read_data("File: Лекция", TestFileOperations.c_sheet)
        self.assertEqual(expected, actual)    

if __name__ == '__main__':
    unittest.main()        